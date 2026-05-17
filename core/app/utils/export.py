"""
Export utilities — produces CSV and Excel (.xlsx) files
Used by /reports endpoints with ?format=csv or ?format=xlsx
"""

import csv
import io
from typing import Any

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi.responses import StreamingResponse


def export_to_csv(headers: list[str], rows: list[list[Any]], filename: str) -> StreamingResponse:
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(headers)
    writer.writerows(rows)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="{filename}.csv"'},
    )


def export_to_xlsx(headers: list[str], rows: list[list[Any]], filename: str, sheet_name: str = "Report") -> StreamingResponse:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row styling
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(header) + 4, 15)

    # Data rows
    for row_idx, row in enumerate(rows, 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return StreamingResponse(
        iter([buffer.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}.xlsx"'},
    )


def build_export_response(
    headers: list[str],
    rows: list[list[Any]],
    filename: str,
    fmt: str = "csv",
    sheet_name: str = "Report",
) -> StreamingResponse:
    if fmt == "xlsx":
        return export_to_xlsx(headers, rows, filename, sheet_name)
    return export_to_csv(headers, rows, filename)
